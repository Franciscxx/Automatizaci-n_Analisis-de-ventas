pip install pandas openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill

# Carga el archivo de Excel
file_path = 'tu_archivo.xlsx'
wb = load_workbook(file_path)
ws = wb.active

# BORRAMOS LAS COLUMNAS INNECESARIAS
for col in ['A', 'B', 'C', 'D', 'E']:
    ws[f'{col}1'].value = None

for col in ['J', 'K', 'L', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
    for cell in ws[col]:
        cell.value = None

# REORDENAMOS LOS DATOS
for idx, cell in enumerate(ws['M']):
    ws[f'J{idx + 1}'].value = cell.value
    cell.value = None

for idx, cell in enumerate(ws['F']):
    ws[f'A{idx + 1}'].value = cell.value

for col in ['F', 'G', 'H', 'I', 'J', 'K']:
    for cell in ws[col]:
        cell.value = None

# NOMBRES Y FORMATO
headers = ["Regular", "Premium", "Diesel", "DryStock", "Ventas Mes", "Total", "Total Sin IVA"]
for col, header in zip(['G', 'H', 'I', 'J', 'K', 'L', 'M'], headers):
    ws[f'{col}1'].value = header

# Formateo de columnas y filas
ws.column_dimensions['E'].width = 15
comma_style = NamedStyle(name="comma")
comma_style.number_format = '0.00'
for col in ws.iter_cols(min_col=7, max_col=13, min_row=2, max_row=2):
    for cell in col:
        cell.style = comma_style

# CALCULO DE TOTALES Y SIN IVA
datos = pd.DataFrame(ws.values).dropna(how="all")
suma_total = suma_ventas = suma_dry_stock = suma_diesel = suma_premium = suma_regular = 0
suma_dry_stock_sin_iva = suma_diesel_sin_iva = suma_premium_sin_iva = suma_regular_sin_iva = 0

for index, row in datos.iterrows():
    if index == 0:
        continue  # Ignorar la fila de encabezados
    suma_total += row[3]
    if row[0] == "BP DIESEL � MX":
        suma_diesel += row[3]
        suma_diesel_sin_iva += (((((row[2]) - 0.43363) / 1.16) + 0.43363) * row[1])
    elif row[0] == "BP PREMIUM 91 O SUPERIOR � MX":
        suma_premium += row[3]
        suma_premium_sin_iva += (((((row[2]) - 0.63752) / 1.16) + 0.63752) * row[1])
    elif row[0] == "BP REGULAR 87 � MX":
        suma_regular += row[3]
        suma_regular_sin_iva += (((((row[2]) - 0.52248) / 1.16) + 0.52248) * row[1])
    else:
        suma_dry_stock_sin_iva += ((row[2] / 1.16) * row[1])

suma_dry_stock = suma_total - suma_ventas - suma_diesel - suma_premium - suma_regular
total_sin_iva = suma_dry_stock_sin_iva + suma_diesel_sin_iva + suma_premium_sin_iva + suma_regular_sin_iva

# RESULTADOS
ws['G2'].value = suma_regular
ws['H2'].value = suma_premium
ws['I2'].value = suma_diesel
ws['J2'].value = suma_dry_stock
ws['K2'].value = suma_ventas
ws['L2'].value = suma_total
ws['M2'].value = total_sin_iva

# FORMATEAR Y DESTACAR RESULTADOS
for col in ['G', 'H', 'I', 'J', 'K', 'L', 'M']:
    ws.column_dimensions[col].width = 15

# Destacar los resultados con un color de fondo
highlight_fill = PatternFill(start_color="51BA9C", end_color="51BA9C", fill_type="solid")
for col in ['G', 'H', 'I', 'J']:
    ws[f'{col}3'].fill = highlight_fill

# Guarda los cambios en el archivo de Excel
wb.save('tu_archivo_modificado.xlsx')